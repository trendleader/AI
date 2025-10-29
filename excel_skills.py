import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import openai
import os
import json
from datetime import datetime
from io import BytesIO
import base64

# Set page configuration
st.set_page_config(
    page_title="Excel Skills Assessment",
    page_icon="📊",
    layout="wide"
)

# Initialize session state
if 'exam_started' not in st.session_state:
    st.session_state.exam_started = False
if 'current_question' not in st.session_state:
    st.session_state.current_question = 0
if 'submitted_files' not in st.session_state:
    st.session_state.submitted_files = {}
if 'evaluations' not in st.session_state:
    st.session_state.evaluations = {}
if 'exam_completed' not in st.session_state:
    st.session_state.exam_completed = False
if 'openai_api_key' not in st.session_state:
    st.session_state.openai_api_key = None
if 'api_key_set' not in st.session_state:
    st.session_state.api_key_set = False

# Excel Tasks Database
EXCEL_TASKS = [
    {
        "id": 1,
        "title": "Sales Tax Calculation",
        "difficulty": "Easy",
        "description": """Your company sells products to customers in two cities: New York and San Francisco.
        
**Task:**
- Download the template file 'SalesTax-Template.xlsx'
- Write formulas to calculate retail prices inclusive of sales tax using base prices
- Must reference cells A5 (NY Tax Rate) and B5 (SF Tax Rate) in your formulas
- Create two columns: "NY Price (with tax)" and "SF Price (with tax)"
- Save and upload your completed file

**Skills Tested:** VLOOKUP, Cell References, Formula Writing""",
        "time_limit": 15,
        "sample_data": {
            "products": ["Laptop", "Mouse", "Keyboard", "Monitor", "Headphones"],
            "base_prices": [999, 25, 75, 350, 150]
        }
    },
    {
        "id": 2,
        "title": "Date Filtering & Data Organization",
        "difficulty": "Medium",
        "description": """Your accounting department needs orders filtered by specific time periods.

**Task:**
- Download 'DateFilter-Template.xlsx'
- On Sheet1: Filter orders from July 1, 2018 through August 31, 2018
- On Sheet2: Filter orders before July 1, 2018 AND after August 31, 2018
- Maintain data integrity and proper sorting
- Save and upload your completed file

**Skills Tested:** Advanced Filtering, Date Functions, Data Organization""",
        "time_limit": 20,
        "sample_data": None
    },
    {
        "id": 3,
        "title": "Pivot Table Analysis",
        "difficulty": "Medium",
        "description": """Analyze sales data using pivot tables.

**Task:**
- Download 'PivotTable-Template.xlsx'
- Create a pivot table summarizing sales by region and product
- Sum the revenue values
- Add region names as rows and product names as columns
- Format the pivot table with number formatting (2 decimal places)
- Save and upload your completed file

**Skills Tested:** Pivot Tables, Data Summarization, Formatting""",
        "time_limit": 20,
        "sample_data": None
    },
    {
        "id": 4,
        "title": "Chart Creation & Customization",
        "difficulty": "Easy",
        "description": """Create and customize a professional chart.

**Task:**
- Download 'Chart-Template.xlsx'
- Create a clustered column chart comparing quarterly sales
- Set chart title: "Q1-Q4 Sales Comparison"
- Add axis titles: "Quarter" (horizontal), "Sales ($)" (vertical)
- Remove gridlines for a clean look
- Add a legend and data labels
- Save and upload your completed file

**Skills Tested:** Charting, Chart Formatting, Data Visualization""",
        "time_limit": 15,
        "sample_data": None
    },
    {
        "id": 5,
        "title": "VLOOKUP & Data Consolidation",
        "difficulty": "Medium",
        "description": """Use VLOOKUP to consolidate employee data.

**Task:**
- Download 'VLOOKUP-Template.xlsx'
- Use VLOOKUP to match employee IDs with their departments
- Create a lookup reference table with Employee ID, Name, Department
- Write VLOOKUP formulas to populate missing department information
- Add error handling (IFERROR)
- Save and upload your completed file

**Skills Tested:** VLOOKUP, Error Handling, Data Consolidation""",
        "time_limit": 20,
        "sample_data": None
    },
    {
        "id": 6,
        "title": "VBA Macros & Automation",
        "difficulty": "Hard",
        "description": """Create VBA macros to automate Excel tasks.

**Task:**
- Download 'Macros-Template.xlsm'
- Create a macro that automatically formats a data range (bold header, alternating row colors)
- Create a second macro that calculates and displays the sum of selected cells in a message box
- Add a button to Sheet1 labeled "Format Data" that runs the formatting macro
- Add a button labeled "Calculate Sum" that runs the sum macro
- Test both macros to ensure they work correctly
- Save as .xlsm format (Excel Macro-Enabled Workbook)
- Upload your completed file

**Macro Requirements:**
- Macro 1: Format header row in bold, apply light blue background to header, alternate row colors (white/light gray)
- Macro 2: Calculate sum of selected cells and display in a message box
- Both macros should include comments explaining the code
- Save file with both macros functional

**Skills Tested:** VBA Programming, Macro Creation, Automation, Event Handling""",
        "time_limit": 25,
        "sample_data": None
    }
]

def get_openai_client():
    """Initialize OpenAI client from session state, secrets, or environment"""
    # Check session state first
    if st.session_state.openai_api_key:
        return openai.OpenAI(api_key=st.session_state.openai_api_key)
    
    # Fall back to secrets
    api_key = st.secrets.get("OPENAI_API_KEY") or os.environ.get("OPENAI_API_KEY")
    if api_key:
        st.session_state.openai_api_key = api_key
        st.session_state.api_key_set = True
        return openai.OpenAI(api_key=api_key)
    
    return None

def create_template_file(task_id: int) -> BytesIO:
    """Create Excel template file based on task"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    if task_id == 1:
        # Sales Tax Template
        ws['A1'] = "Product"
        ws['B1'] = "Base Price"
        ws['A5'] = "NY Tax Rate (%)"
        ws['B5'] = 8.875
        ws['A6'] = "SF Tax Rate (%)"
        ws['B6'] = 8.625
        
        products = ["Laptop", "Mouse", "Keyboard", "Monitor", "Headphones"]
        prices = [999, 25, 75, 350, 150]
        
        for idx, (prod, price) in enumerate(zip(products, prices), start=2):
            ws[f'A{idx}'] = prod
            ws[f'B{idx}'] = price
        
        ws['C1'] = "NY Price (with tax)"
        ws['D1'] = "SF Price (with tax)"
        
    elif task_id == 2:
        # Date Filter Template
        ws['A1'] = "Order ID"
        ws['B1'] = "Order Date"
        ws['C1'] = "Amount"
        ws['D1'] = "Customer"
        
        dates = [
            "2018-06-15", "2018-07-10", "2018-08-20", "2018-09-05",
            "2018-07-25", "2018-06-01", "2018-08-31", "2018-09-15"
        ]
        
        for idx, date in enumerate(dates, start=2):
            ws[f'A{idx}'] = f"ORD{idx-1:04d}"
            ws[f'B{idx}'] = date
            ws[f'C{idx}'] = 1000 + (idx * 100)
            ws[f'D{idx}'] = f"Customer {idx-1}"
    
    elif task_id == 3:
        # Pivot Table Template
        ws['A1'] = "Region"
        ws['B1'] = "Product"
        ws['C1'] = "Revenue"
        
        data = [
            ("North", "Widget", 5000),
            ("South", "Widget", 4500),
            ("North", "Gadget", 6000),
            ("South", "Gadget", 5500),
            ("East", "Widget", 4200),
            ("East", "Gadget", 5800),
        ]
        
        for idx, (region, product, revenue) in enumerate(data, start=2):
            ws[f'A{idx}'] = region
            ws[f'B{idx}'] = product
            ws[f'C{idx}'] = revenue
    
    elif task_id == 4:
        # Chart Template
        ws['A1'] = "Quarter"
        ws['B1'] = "Sales"
        
        quarters = ["Q1", "Q2", "Q3", "Q4"]
        sales = [45000, 52000, 48500, 61000]
        
        for idx, (quarter, sale) in enumerate(zip(quarters, sales), start=2):
            ws[f'A{idx}'] = quarter
            ws[f'B{idx}'] = sale
    
    elif task_id == 5:
        # VLOOKUP Template
        ws['A1'] = "Employee ID"
        ws['B1'] = "Name"
        ws['C1'] = "Department"
        
        # Reference data
        ref_data = [
            (101, "John Smith", "Sales"),
            (102, "Jane Doe", "Marketing"),
            (103, "Bob Johnson", "IT"),
        ]
        
        for idx, (emp_id, name, dept) in enumerate(ref_data, start=2):
            ws[f'A{idx}'] = emp_id
            ws[f'B{idx}'] = name
            ws[f'C{idx}'] = dept
        
        # Employee data to lookup
        ws['E1'] = "Lookup Data"
        ws['E2'] = "Employee ID"
        ws['F2'] = "Department (to find)"
        
        lookup_ids = [101, 102, 103, 104]
        for idx, emp_id in enumerate(lookup_ids, start=3):
            ws[f'E{idx}'] = emp_id
    
    elif task_id == 6:
        # VBA Macros Template
        ws['A1'] = "Product"
        ws['B1'] = "Q1 Sales"
        ws['C1'] = "Q2 Sales"
        ws['D1'] = "Q3 Sales"
        ws['E1'] = "Q4 Sales"
        
        data = [
            ("Widget", 1000, 1200, 1100, 1300),
            ("Gadget", 800, 950, 1050, 1200),
            ("Doohickey", 600, 700, 750, 850),
            ("Thingamabob", 500, 600, 650, 750),
        ]
        
        for idx, (product, q1, q2, q3, q4) in enumerate(data, start=2):
            ws[f'A{idx}'] = product
            ws[f'B{idx}'] = q1
            ws[f'C{idx}'] = q2
            ws[f'D{idx}'] = q3
            ws[f'E{idx}'] = q4
        
        # Add instructions
        ws['A10'] = "Instructions:"
        ws['A11'] = "1. Create a macro 'FormatData' to format this table"
        ws['A12'] = "2. Create a macro 'CalculateSum' to sum selected cells"
        ws['A13'] = "3. Add buttons to run these macros"
        ws['A14'] = "4. Save as .xlsm (macro-enabled)"
        
        # Macro VBA code as reference in comment
        ws['A16'] = "Sample VBA Code Reference:"
        ws['A17'] = "Sub FormatData()"
        ws['A18'] = "  'Format header row and alternate colors"
        ws['A19'] = "End Sub"
    
    # Format headers
    for cell in ws[1]:
        if cell.value:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def analyze_submitted_file(task_id: int, uploaded_file) -> dict:
    """Analyze submitted Excel file"""
    try:
        # Read the uploaded file
        df = pd.read_excel(uploaded_file)
        
        # Basic validation
        analysis = {
            "file_name": uploaded_file.name,
            "sheet_count": 0,
            "row_count": len(df),
            "column_count": len(df.columns),
            "has_data": not df.empty,
            "columns": list(df.columns),
            "preview": df.head().to_dict('records')
        }
        
        # Task-specific analysis
        if task_id == 1:
            analysis["has_tax_columns"] = any("tax" in str(col).lower() for col in df.columns)
            analysis["has_formulas"] = "formula_check_needed"
        
        elif task_id == 2:
            analysis["sheet_count"] = len(pd.ExcelFile(uploaded_file).sheet_names)
            analysis["has_multiple_sheets"] = analysis["sheet_count"] > 1
        
        elif task_id == 3:
            analysis["has_pivot_table"] = "pivot" in str(df.columns).lower()
        
        elif task_id == 4:
            analysis["data_present"] = not df.empty
        
        elif task_id == 5:
            analysis["has_lookup_formulas"] = any("vlookup" in str(val).lower() for val in df.values.flatten() if isinstance(val, str))
        
        return analysis
        
    except Exception as e:
        return {
            "error": str(e),
            "file_name": uploaded_file.name if hasattr(uploaded_file, 'name') else "unknown"
        }

def evaluate_submission(task_id: int, file_analysis: dict) -> dict:
    """Use OpenAI to evaluate the submission"""
    client = get_openai_client()
    if not client:
        return {
            "score": 0,
            "feedback": "API key not configured",
            "error": True
        }
    
    task = EXCEL_TASKS[task_id - 1]
    
    evaluation_prompt = f"""You are an expert Excel instructor evaluating a student's Excel assignment.

Task: {task['title']}
Description: {task['description']}
Difficulty: {task['difficulty']}

Student's Submission Analysis:
{json.dumps(file_analysis, indent=2)}

Based on the submission analysis provided, evaluate how well the student completed the Excel task. Consider:

1. **Completeness** - Did they complete all required parts?
2. **Data Integrity** - Is the data organized and correct?
3. **Formula Accuracy** - Are formulas written correctly (if applicable)?
4. **Formatting** - Is the spreadsheet well-formatted and professional?
5. **Excel Skills** - Does the work demonstrate competence with Excel features?

Provide your evaluation in JSON format:
{{
    "score": 85,
    "completeness": "description of what was completed",
    "strengths": ["strength1", "strength2"],
    "improvements": ["area1", "area2"],
    "feedback": "overall feedback on the submission",
    "excel_skills_demonstrated": ["skill1", "skill2"]
}}

Only respond with valid JSON, no additional text."""

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "user", "content": evaluation_prompt}
            ],
            temperature=0.7
        )
        
        response_text = response.choices[0].message.content
        
        try:
            start_idx = response_text.find('{')
            end_idx = response_text.rfind('}') + 1
            json_str = response_text[start_idx:end_idx]
            evaluation = json.loads(json_str)
            return evaluation
        except json.JSONDecodeError:
            return {
                "score": 0,
                "feedback": response_text,
                "error": True
            }
            
    except Exception as e:
        return {
            "score": 0,
            "feedback": f"Error evaluating submission: {str(e)}",
            "error": True
        }

def display_api_key_setup():
    """Display API key input interface"""
    st.title("🔐 OpenAI API Key Configuration")
    st.markdown("---")
    
    st.info("""
    To use this Excel Assessment app, you need to provide your OpenAI API key.
    
    **How to get an API key:**
    1. Go to https://platform.openai.com/api/keys
    2. Sign in with your OpenAI account (create one if needed)
    3. Click "Create new secret key"
    4. Copy the key and paste it below
    
    **Security Note:** Your API key is stored securely in this session only and will not be saved or transmitted to any server.
    """)
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        api_key_input = st.text_input(
            "Enter your OpenAI API Key:",
            type="password",
            placeholder="sk-...",
            help="Your API key will be used to evaluate your Excel submissions"
        )
    
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("✅ Set API Key", use_container_width=True, type="primary"):
            if api_key_input and api_key_input.startswith("sk-"):
                st.session_state.openai_api_key = api_key_input
                st.session_state.api_key_set = True
                st.success("✅ API key configured successfully!")
                st.rerun()
            else:
                st.error("❌ Invalid API key format. Please check and try again.")
    
    st.markdown("---")
    
    # Alternative: Check if key is already set via secrets or environment
    if not st.session_state.api_key_set:
        client = get_openai_client()
        if client:
            st.success("✅ API key detected from environment/secrets. Click 'Start Assessment' below to proceed.")
            st.session_state.api_key_set = True
            return True
    
    return st.session_state.api_key_set

def display_welcome_screen():
    """Display welcome and instructions"""
    st.title("📊 Excel Skills Assessment")
    st.markdown("---")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.header("Test Your Excel Proficiency")
        st.write("""
        This assessment evaluates your practical Excel skills through real-world tasks.
        Download templates, complete the tasks in Excel, and upload your results for instant evaluation.
        
        **Assessment Details:**
        - Total Tasks: 6
        - Difficulty Range: Easy to Hard
        - Time Recommendations: 15-25 minutes per task
        - Format: Download templates → Complete in Excel → Upload files
        
        **What You'll Do:**
        1. Download an Excel template for each task
        2. Complete the task using Excel (formulas, charts, filtering, macros, etc.)
        3. Upload your completed file
        4. Receive instant AI-powered evaluation with detailed feedback
        5. View your overall performance score
        """)
    
    with col2:
        st.info("""
        **Excel Features Tested:**
        ✓ Formulas & Functions
        ✓ Data Filtering
        ✓ Pivot Tables
        ✓ Chart Creation
        ✓ VLOOKUP & Lookups
        ✓ Data Formatting
        ✓ Error Handling
        ✓ VBA Macros
        """)
    
    st.markdown("---")
    
    if st.button("🚀 Start Assessment", key="start_exam", use_container_width=True, type="primary"):
        st.session_state.exam_started = True
        st.rerun()

def display_task_screen():
    """Display the current task"""
    task = EXCEL_TASKS[st.session_state.current_question]
    
    # Progress bar
    progress = st.session_state.current_question / len(EXCEL_TASKS)
    st.progress(progress, text=f"Task {st.session_state.current_question + 1} of {len(EXCEL_TASKS)}")
    
    # Task header
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        st.title(f"Task {st.session_state.current_question + 1}: {task['title']}")
    with col2:
        difficulty_color = "🟢" if task['difficulty'] == "Easy" else "🟡"
        st.metric("Difficulty", f"{difficulty_color} {task['difficulty']}")
    with col3:
        st.metric("Time Limit", f"{task['time_limit']} min")
    
    st.markdown("---")
    
    # Task description
    st.subheader("📋 Task Description")
    st.write(task['description'])
    
    st.markdown("---")
    
    # Download template
    st.subheader("📥 Download Template")
    
    template_file = create_template_file(task['id'])
    
    # Use .xlsm for macros task, .xlsx for others
    file_ext = ".xlsm" if task['id'] == 6 else ".xlsx"
    file_name = f"Task{task['id']}-Template{file_ext}"
    
    st.download_button(
        label=f"⬇️ Download Excel Template",
        data=template_file,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
    
    st.info("""
    **Steps:**
    1. Download the Excel template above
    2. Open it in Microsoft Excel or Excel Online
    3. Complete all required tasks
    4. Save the file in .xlsx format
    5. Upload your completed file below
    """)
    
    st.markdown("---")
    
    # File upload
    st.subheader("📤 Upload Completed File")
    
    uploaded_file = st.file_uploader(
        "Upload your completed Excel file:",
        type=["xlsx", "xls"],
        key=f"upload_{task['id']}"
    )
    
    if uploaded_file is not None:
        st.success(f"✅ File uploaded: {uploaded_file.name}")
        
        # Show file preview
        try:
            df_preview = pd.read_excel(uploaded_file)
            st.subheader("File Preview")
            st.dataframe(df_preview.head(10), use_container_width=True)
        except Exception as e:
            st.warning(f"Could not preview file: {str(e)}")
    
    st.markdown("---")
    
    # Navigation and submission
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        if st.session_state.current_question > 0:
            if st.button("⬅️ Previous", use_container_width=True):
                st.session_state.current_question -= 1
                st.rerun()
    
    with col2:
        if st.button("⏭️ Skip", use_container_width=True):
            if st.session_state.current_question < len(EXCEL_TASKS) - 1:
                st.session_state.current_question += 1
                st.rerun()
            else:
                st.info("You've reached the last task. Click 'Finish Assessment' to complete.")
    
    with col3:
        if st.button("✅ Submit & Evaluate", use_container_width=True, type="primary"):
            if uploaded_file is not None:
                with st.spinner("Analyzing your Excel file..."):
                    file_analysis = analyze_submitted_file(task['id'], uploaded_file)
                    st.session_state.submitted_files[task['id']] = uploaded_file.name
                
                with st.spinner("Evaluating your submission with AI..."):
                    evaluation = evaluate_submission(task['id'], file_analysis)
                    st.session_state.evaluations[task['id']] = evaluation
                
                st.rerun()
            else:
                st.warning("Please upload a file before submitting.")
    
    with col4:
        if st.button("⏹️ End Assessment", use_container_width=True):
            st.session_state.exam_completed = True
            st.rerun()
    
    # Show evaluation if available
    if task['id'] in st.session_state.evaluations:
        st.markdown("---")
        st.subheader("📊 Evaluation Result")
        
        evaluation = st.session_state.evaluations[task['id']]
        
        if evaluation.get("error"):
            st.error(evaluation.get("feedback", "Error during evaluation"))
        else:
            # Score display
            score = evaluation.get("score", 0)
            col1, col2, col3 = st.columns(3)
            col1.metric("Score", f"{score}/100")
            
            # Feedback tabs
            tab1, tab2, tab3, tab4 = st.tabs(["Overview", "Feedback", "Strengths", "Improvements"])
            
            with tab1:
                st.write(evaluation.get("feedback", "No feedback available"))
                if evaluation.get("excel_skills_demonstrated"):
                    st.write("**Excel Skills Demonstrated:**")
                    for skill in evaluation.get("excel_skills_demonstrated", []):
                        st.write(f"✓ {skill}")
            
            with tab2:
                st.write(evaluation.get("completeness", "Not evaluated"))
            
            with tab3:
                if evaluation.get("strengths"):
                    for strength in evaluation.get("strengths", []):
                        st.write(f"✅ {strength}")
            
            with tab4:
                if evaluation.get("improvements"):
                    for improvement in evaluation.get("improvements", []):
                        st.write(f"💡 {improvement}")

def display_results_screen():
    """Display final assessment results"""
    st.title("📊 Assessment Results")
    st.markdown("---")
    
    # Calculate overall score
    scores = []
    for task_id, evaluation in st.session_state.evaluations.items():
        if not evaluation.get("error"):
            scores.append(evaluation.get("score", 0))
    
    if scores:
        overall_score = sum(scores) / len(scores)
        tasks_evaluated = len(scores)
    else:
        overall_score = 0
        tasks_evaluated = 0
    
    tasks_total = len(EXCEL_TASKS)
    
    # Summary metrics
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Overall Score", f"{overall_score:.1f}/100")
    col2.metric("Tasks Completed", f"{tasks_evaluated}/{tasks_total}")
    col3.metric("Average Score", f"{overall_score:.1f}")
    col4.metric("Completion", f"{(tasks_evaluated/tasks_total)*100:.0f}%")
    
    # Performance level
    if overall_score >= 85:
        st.success("🌟 Excellent Performance! You demonstrated strong Excel skills.")
    elif overall_score >= 70:
        st.info("👍 Good Performance. You have solid Excel skills with room for improvement.")
    elif overall_score >= 50:
        st.warning("📈 Fair Performance. Consider practicing more Excel features.")
    else:
        st.error("⚠️ Please review Excel fundamentals and retake the assessment.")
    
    st.markdown("---")
    
    # Detailed results table
    st.subheader("📋 Task Results")
    
    results_data = []
    for task in EXCEL_TASKS:
        task_id = task['id']
        if task_id in st.session_state.evaluations:
            evaluation = st.session_state.evaluations[task_id]
            if not evaluation.get("error"):
                results_data.append({
                    "Task": task['title'],
                    "Difficulty": task['difficulty'],
                    "Score": f"{evaluation.get('score', 0)}/100",
                    "File": st.session_state.submitted_files.get(task_id, "N/A"),
                    "Status": "✅ Evaluated"
                })
            else:
                results_data.append({
                    "Task": task['title'],
                    "Difficulty": task['difficulty'],
                    "Score": "0/100",
                    "File": st.session_state.submitted_files.get(task_id, "N/A"),
                    "Status": "❌ Error"
                })
        else:
            results_data.append({
                "Task": task['title'],
                "Difficulty": task['difficulty'],
                "Score": "0/100",
                "File": "Not submitted",
                "Status": "⏭️ Skipped"
            })
    
    df_results = pd.DataFrame(results_data)
    st.dataframe(df_results, use_container_width=True, hide_index=True)
    
    st.markdown("---")
    
    # Detailed feedback
    st.subheader("📝 Detailed Feedback by Task")
    
    for task in EXCEL_TASKS:
        if task['id'] in st.session_state.evaluations:
            evaluation = st.session_state.evaluations[task['id']]
            
            with st.expander(f"Task {task['id']}: {task['title']} - Score: {evaluation.get('score', 0)}/100"):
                if not evaluation.get("error"):
                    col1, col2 = st.columns([1, 1])
                    
                    with col1:
                        st.subheader("Score Breakdown")
                        st.metric("Score", f"{evaluation.get('score', 0)}/100")
                        st.write(f"**Completeness:** {evaluation.get('completeness', 'N/A')}")
                    
                    with col2:
                        st.subheader("Feedback")
                        st.write(evaluation.get('feedback', 'No feedback available'))
                    
                    if evaluation.get("strengths"):
                        st.write("**Strengths:**")
                        for strength in evaluation.get("strengths", []):
                            st.write(f"✅ {strength}")
                    
                    if evaluation.get("improvements"):
                        st.write("**Areas for Improvement:**")
                        for improvement in evaluation.get("improvements", []):
                            st.write(f"💡 {improvement}")
                else:
                    st.error(evaluation.get("feedback", "Error during evaluation"))
    
    st.markdown("---")
    
    # Export results
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("📥 Download Results as CSV", use_container_width=True):
            csv = df_results.to_csv(index=False)
            st.download_button(
                label="Download CSV",
                data=csv,
                file_name=f"assessment_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )
    
    with col2:
        if st.button("🔄 Retake Assessment", use_container_width=True):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

def main():
    """Main app logic"""
    
    # Sidebar
    with st.sidebar:
        st.title("📊 Assessment Info")
        st.info(f"""
        **Status:** {'🟢 In Progress' if st.session_state.exam_started else '🟡 Not Started'}
        
        **Tasks:** {len(EXCEL_TASKS)}
        
        **Difficulty:** Easy - Hard
        """)
        
        if st.session_state.exam_started and not st.session_state.exam_completed:
            st.markdown("---")
            st.subheader("Tasks Overview")
            for i, task in enumerate(EXCEL_TASKS):
                status = "✅" if task['id'] in st.session_state.evaluations else "📝"
                if st.button(f"{status} Task {i+1}: {task['title']}", use_container_width=True):
                    st.session_state.current_question = i
                    st.rerun()
        
        if st.session_state.api_key_set:
            st.markdown("---")
            st.success("✅ API Key Configured")
            if st.button("🔄 Change API Key"):
                st.session_state.api_key_set = False
                st.session_state.openai_api_key = None
                st.rerun()
    
    # Main content
    if not st.session_state.api_key_set:
        display_api_key_setup()
    elif not st.session_state.exam_started:
        display_welcome_screen()
    elif st.session_state.exam_completed:
        display_results_screen()
    else:
        display_task_screen()

if __name__ == "__main__":
    main()
